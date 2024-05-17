import os
import json, time
from loguru import logger
import requests
from openpyxl import Workbook
from datetime import datetime
import openpyxl
'''Settings use or not use proxies and filename'''
EXEL_FILENAME = 'scroll_marks.xlsx'
USE_PROXY = False

with open("addresses.txt", "r") as f:
    WALLETS = [row.strip() for row in f]

with open("proxies.txt", "r") as f:
    PROXIES = [row.strip() for row in f]

def get_wallet_proxies(wallets, proxies):
    try:
        result = {}
        for i in range(len(wallets)):
            result[wallets[i]] = proxies[i % len(proxies)]
        return result
    except: None

WALLET_PROXIES  = get_wallet_proxies(WALLETS, PROXIES)


def request(method="get", request_retry=0, wallet=0, proxy=None, **kwargs):
    session = requests.Session()

    if proxy is not None:
        session.proxies.update(
            {
                "http": f"{proxy}",
                "https": f"{proxy}"
            }
        )

    if request_retry > 4:
        return
    retry = 0
    while True:
        try:
            if method == "post":
                response = session.post(**kwargs, verify=False)
            elif method == "get":
                response = session.get(**kwargs, verify=False)
            elif method == "put":
                response = session.put(**kwargs, verify=False)
            elif method == "options":
                response = session.options(**kwargs, verify=False)

            logger.info(f'{wallet}, status_code {response.status_code} response: {response}')

            if response.status_code == 201 or response.status_code == 200:
                time.sleep(5)
                try:
                    return response.json()
                except json.decoder.JSONDecodeError:
                    logger.info('The request success but not contain a JSON')
                    break
            else:
                logger.error(f'[{wallet} - Bad status code: {response.status_code} {response.json()}')
                time.sleep(15)
                retry += 1
                if retry > 4:
                    break

        except Exception as error:
            logger.error(f'{wallet} - {kwargs["url"]} failed to make request | {error}')
            time.sleep(15)
            request(method=method, request_retry=request_retry + 1, wallet=wallet, proxy=proxy, **kwargs)
            break

def get_transaction_list(wallet, chain):
    url = f'https://kx58j6x5me.execute-api.us-east-1.amazonaws.com/scroll/bridge-balances?walletAddress={wallet}'

    headers = {
        'accept': 'application/json',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
    }
    
    response = requests.get(url=url, headers=headers)
    print(f'get_transactions_count {response.status_code}')

    if response.status_code == 200:
        try:
            data = response.json()
            return data

        except ValueError as e:
            print("Error decoding JSON:", e)
            print("Response content:", response.content)
    else:
        print('Error')

def set_column_widths(sheet):
    column_widths = {
        'A': 20,  # Date Time
        'B': 50,  # Wallet
        'C': 15,  # Amount
        'D': 15,  # Points
        'E': 15,  # Value in USD
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

def add_data(wallet: str, amount: float, points: int, value_in_usd: float):
    file_name = EXEL_FILENAME
    if not os.path.exists(file_name):
        book = Workbook()
        sheet = book.active
        sheet['A1'] = 'Date Time'
        sheet['B1'] = 'Wallet'
        sheet['C1'] = 'Amount'
        sheet['D1'] = 'Points'
        sheet['E1'] = 'Value in USD'
        set_column_widths(sheet)
        book.save(file_name)
        book.close()

    book = openpyxl.load_workbook(file_name)
    sheet = book.active

    new_row = sheet.max_row + 1
    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.cell(row=new_row, column=1).value = current_datetime
    sheet.cell(row=new_row, column=2).value = wallet
    sheet.cell(row=new_row, column=3).value = '{:.10f}'.format(amount)
    sheet.cell(row=new_row, column=4).value = '{:.10f}'.format(points)
    sheet.cell(row=new_row, column=5).value = '{:.10f}'.format(value_in_usd)

    book.save(file_name)
    book.close()
    logger.success(f'Data added to {file_name}: {current_datetime}, {wallet}, {amount}, {points}, {value_in_usd}')

def start_check_marks(wallet):
    proxy = None
    if USE_PROXY:
        proxy = WALLET_PROXIES[wallet]

    url = f'https://kx58j6x5me.execute-api.us-east-1.amazonaws.com/scroll/bridge-balances?walletAddress={wallet}'

    response = request(url=url, wallet=wallet, proxy=proxy)

    amount = 0
    value_in_usd = 0
    points = 0

    for item in response:
        amount = amount + item.get("amount")
        value_in_usd = value_in_usd + item.get("value_in_usd")
        points = points + item.get("points")

    add_data(wallet, amount, points, value_in_usd)

if __name__ == '__main__':
    for wallet in WALLETS:
        start_check_marks(wallet)
    logger.info('Verified all data')